from urllib2 import urlopen as uReq
import requests
from bs4 import BeautifulSoup as soup
import xlrd
import openpyxl
import os
import string

#dir = r'C:\Users\Tien'
#os.chdir(dir)

########### input excel ##########
input_workbook = xlrd.open_workbook("aktie.xlsx")

input_worksheet = input_workbook.sheet_by_index(0)



############### output excel ####################
rawdata_workbook = openpyxl.load_workbook('rawdata.xlsx')

rawdata_worksheet = rawdata_workbook.get_sheet_by_name('Sheet1')


################### global variables ####################
i = 1141
alphabet = list(string.ascii_uppercase)


while i <= 1854:
	try:
		#stockname = input_worksheet.cell(i,0).value
		i = i + 1 
		

		#googleurl

		#stockname = stockname.replace(" ","+")



		#search_url = 'http://www.finanzen.net/suchergebnis.asp?_search=' + stockname 

		#print(search_url)

		#search_uClient = uReq(search_url)
		#search_page_html = search_uClient.read()
		#search_uClient.close()

		#page_soup = soup(search_page_html, "html.parser")

		#grab the box with stock information

		#my_url = page_soup.findAll("table",{"class":"table"})

		#finanzen_stockname = my_url[0].contents[2].a.attrs['href'].replace("/aktien/","")

		#my_url = 'http://www.finanzen.net' + my_url[0].contents[2].a.attrs['href']

		my_url = input_worksheet.cell(i,0).value

		finanzen_stockname = my_url.replace("http://www.finanzen.net/aktien/","")

		print(my_url)

		print('http://www.finanzen.net/bilanz_guv/' + finanzen_stockname.replace("-Aktie",""))

		guv_url = 'http://www.finanzen.net/bilanz_guv/' + finanzen_stockname.replace("-Aktie","")


		###########################################   MAIN SITE  http://www.finanzen.net/aktien/BASF-Aktie #############################################

		#url reader function
		uClient = uReq(my_url)
		page_html = uClient.read()
		uClient.close()

		page_soup = soup(page_html, "html.parser")

		#grab isin
		try:
			isin = page_soup.findAll("span",{"class":"instrument-id"})

			isin = isin[0].text

			if isin[0:4] == "ISIN":
				isin = isin[6:17]
			else:
				isin = isin[20:32]

			isin_index = 'A' + str(i-47)
			rawdata_worksheet[isin_index] = isin

		except Exception as e:
			print('no ISIN')


		#grab name

		try:
			stock_index = 'B' + str(i-47)
			rawdata_worksheet[stock_index] = finanzen_stockname.replace("-Aktie","")

		except Exception as e:
			print('no stockname')

		#grab newest stock price

		try:
			stock = page_soup.findAll("div",{"class":"col-xs-5 col-sm-4 text-sm-right text-nowrap"})

			eur_price = stock[0].contents[0]

			eur_price_index = 'C' + str(i-47)
			rawdata_worksheet[eur_price_index] = eur_price

		except Exception as e:
			print('no stockdata')

		
		#grab market capitalization
		try:
			data = page_soup.findAll("div",{"class":"box table-quotes"})

			market_capitalization = data[0].contents[3].contents[5].contents[1].text

			if market_capitalization.find("Mio.") >= 1:
				market_capitalization = market_capitalization.replace(" Mio.","")
				market_capitalization = market_capitalization.replace(",",".")
				market_capitalization = float (market_capitalization)
				market_capitalization = market_capitalization / 10000
				market_capitalization = str(market_capitalization)

			else:
				market_capitalization = market_capitalization.replace(" Mrd.","")
				market_capitalization = market_capitalization.replace(",",".")
				market_capitalization = float (market_capitalization)
				market_capitalization = str(market_capitalization)

			market_capitalization_index = 'D' + str(i-47)
			rawdata_worksheet[market_capitalization_index] = market_capitalization
		except Exception as e:
			print('no market capitalization')

		#grab moody score
		try:
			moody = page_soup.findAll("div",{"class":"tachoValue tachoMcrs mr1"})
			moody = moody[0].text	
			moody_index = 'E' + str(i-47)
			rawdata_worksheet[moody_index] = moody
		except Exception as e:
			print('no moody')

		rawdata_workbook.save('rawdata.xlsx')


		###########################################   BILANZ GUV   #############################################


		uClient = uReq(guv_url)
		page_html = uClient.read()
		uClient.close()

		page_soup = soup(page_html, "html.parser")
		guv = page_soup.findAll("td",{"class":"font-bold"})
		jahr = (['','','','','','',''])

		############################################  check year    ############################################
		year = page_soup.findAll("table",{"class":"table"})
		print(year[1].parent.contents[0].text[5:33])
		if year[1].parent.contents[0].text[29:33] == "2017":
			k = 7
			l = 1
		else:
			k = 8
			l = 0


		##############################################  check currency  ########################################
		try:
			currency = page_soup.findAll("h2",{"class":"box-headline"})
			currency = currency[0].text.split('(in ')
			currency = currency[1]
			currency = currency.replace(')','')
			currency_index = 'AW' + str(i-47)
			rawdata_worksheet[currency_index] = currency
		except Exception as e:
			print ('no currency')


		#########################################     grab kgv    ##############################################
		try:
			j = 2
			while j <= k:
				jahr[j-2] = guv[5].parent.contents[j].text
				kgvindex = alphabet[j+3+l] + str(i-47)
				rawdata_worksheet[kgvindex] = jahr[j-2] 
				j = j+1	
			if l == 1:
				rawdata_worksheet[alphabet[5] + str(i-47)] = '-'
		except Exception as e:
			print ('no kgv')

		#########################################     grab eps    ##############################################
		try:
			j = 2
			while j <= k:
				jahr[j-2] = guv[26].parent.contents[j].text
				epsindex = alphabet[j+10+l] + str(i-47)
				rawdata_worksheet[epsindex] = jahr[j-2]
				j = j+1
			if l == 1:
				rawdata_worksheet[alphabet[12] + str(i-47)] = '-'
		except Exception as e:
			print ('no eps')


		##########################################    grab dividend per share ##################################

		try:
			j = 2
			while j <= k:
				jahr[j-2] = guv[2].parent.contents[j].text
				divindex = alphabet[j+17+l] + str(i-47)
				rawdata_worksheet[divindex] = jahr[j-2]
				j = j+1
			if l == 1:
				rawdata_worksheet[alphabet[19] + str(i-47)] = '-'
		except Exception as e:
			print ('no dividends')


		##########################################    grab own capital   ########################################
		try:
			j = 2
			while j <= k:
				ocap_index = 'A'+ alphabet[j-1+l] + str(i-47)
				rawdata_worksheet[ocap_index] = guv[22].parent.contents[j].text
				j = j+1
			if l == 1:
				rawdata_worksheet['AB' + str(i-47)] = '-'
		except Exception as e:
			print ('no own capital')


		##########################################    grab foreign capital   #####################################
		try:
			j = 2
			while j <= k:
				jahr[j-2] = guv[8].parent.contents[j].text
				fcap_index = 'A'+ alphabet[j+6+l] + str(i-47)
				rawdata_worksheet[fcap_index] = jahr[j-2]
				j = j + 1
			if l == 1:
				rawdata_worksheet['AI' + str(i-47)] = '-'
		except Exception as e:
			print ('no foreign capital')


		##########################################    grab return with taxes   ###################################
		try:
			j = 2
			while j <= k:
				jahr[j-2] = guv[18].parent.contents[j].text
				eps_tax_index = 'A'+ alphabet[j+13+l] + str(i-47)
				rawdata_worksheet[eps_tax_index] = jahr[j-2]
				j = j + 1
			if l == 1:
				rawdata_worksheet['AP' + str(i-47)] = '-'
		except Exception as e:
			print ('no return with taxes')


		rawdata_workbook.save('rawdata.xlsx')


		###################################### Boerse.de      ##############################################

		#########################################      grab price March 2009 ####################################

		try:
			boerse_url = 'http://www.boerse.de/historische-kurse/wertpapier/'+ isin + '_jahr,2009#jahr'
			uClient = uReq(boerse_url)
			page_html = uClient.read()
			uClient.close()
			p2009 = soup(page_html, "html.parser")
			price2009 = p2009.findAll("td",{"class":"alignright"})
			price2009 = price2009[159].text.replace(" ","")
			price2009 = price2009.replace("\n","")
			price2009index = 'AA' + str(i-47)
			rawdata_worksheet[price2009index] = price2009
		except Exception as e:
			print('no 2009 data')



		#########################################      grab price Jan 2017 ######################################
		try:
			boerse_url = 'http://www.boerse.de/historische-kurse/wertpapier/'+ isin + '_jahr,2017#jahr'
			uClient = uReq(boerse_url)
			page_html = uClient.read()
			uClient.close()
			p2017 = soup(page_html, "html.parser")
			price2017 = p2017.findAll("td",{"class":"alignright"})
			price2017 = price2017[139].text.replace(" ","")
			price2017 = price2017.replace("\n","")
			price2017index = 'AX' + str(i-47)
			rawdata_worksheet[price2017index] = price2017
		except Exception as e:
			print('no Jan 2017 data')

		rawdata_workbook.save('rawdata.xlsx')

	except Exception as e:
		print('error')

	


