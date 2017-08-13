import openpyxl
import string


#################### this script substract companies, which are not in both excel files ################

################################### read the current company table 		   #############################

comp_workbook = openpyxl.load_workbook("Aktuelle Unternehmen.xlsx")

comp_worksheet = comp_workbook.active

################################### read the raw data table ############################################

raw_workbook = openpyxl.load_workbook("Aktie.xlsx")

raw_worksheet = raw_workbook.active


##################################### find out the numbers of rows ######################################
comp_length = comp_worksheet.max_row

raw_length = raw_worksheet.max_row

i = 2
j = 2

alphabet = list(string.ascii_uppercase)

while i <= raw_length:
	
	j = 2
	while j <= comp_length:

		if  raw_worksheet['B'+str(i)].value == comp_worksheet['A'+ str(j)].value:
			print(str(i)+' '+str(j))
			raw_worksheet['C'+str(i)].value = '1'
		j = j + 1
			
	i = i + 1 

	
		

raw_workbook.save('Aktie.xlsx')	 