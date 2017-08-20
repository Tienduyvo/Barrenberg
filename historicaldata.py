from urllib.request import urlopen as uReq
import requests
from bs4 import BeautifulSoup as soup
import xlrd
import openpyxl
import os
import string
from selenium import webdriver
import time

#dir = r'C:\Users\Tien'
#os.chdir(dir)

########### input excel ##########
input_workbook = xlrd.open_workbook("aktie.xlsx")

input_worksheet = input_workbook.sheet_by_index(0)



############### output excel ####################
rawdata_workbook = openpyxl.load_workbook('History.xlsx')

rawdata_worksheet = rawdata_workbook.get_sheet_by_name('Tabelle1')


################### global variables ####################
################### index - 2 ###########################
i = 0
alphabet = list(string.ascii_uppercase)




while i <= 630:
	try:

		browser = webdriver.Firefox()
		
		i = i + 1 
		
		my_url = input_worksheet.cell(i,0).value

		print(my_url)

		finanzen_stockname = my_url.replace("http://www.finanzen.net/aktien/","")


		print('http://www.finanzen.net/bilanz_guv/' + finanzen_stockname.replace("-Aktie",""))

		guv_url = 'http://www.finanzen.net/bilanz_guv/' + finanzen_stockname.replace("-Aktie","")


		###########################################   MAIN SITE  http://www.finanzen.net/aktien/BASF-Aktie #############################################

		#url reader function
		uClient = uReq(my_url)
		page_html = uClient.read()
		uClient.close()

		page_soup = soup(page_html, "html.parser")

		#grab isin
		try:
			isin = page_soup.findAll("span",{"class":"instrument-id"})

			isin = isin[0].text

			if isin[0:4] == "ISIN":
				isin = isin[6:17]
			else:
				isin = isin[20:32]

			isin_index = 'A' + str(i)

			rawdata_worksheet[isin_index] = isin

		except Exception as e:
			print('no ISIN')


		#grab name

		try:
			stock_index = 'B' + str(i)
			rawdata_worksheet[stock_index] = finanzen_stockname.replace("-Aktie","")

		except Exception as e:
			print('no stockname')

		#grab newest stock price

		try:
			stock = page_soup.findAll("div",{"col-xs-5 col-sm-4 text-sm-right text-nowrap"})

			eur_price = stock[0].contents[0]

			eur_price_index = 'C' + str(i)
			rawdata_worksheet[eur_price_index] = eur_price

		except Exception as e:
			print('no stockdata')



		###########################################   BILANZ GUV   #############################################


		uClient = uReq(guv_url)
		page_html = uClient.read()
		uClient.close()

		page_soup = soup(page_html, "html.parser")
		guv = page_soup.findAll("td",{"class":"font-bold"})


		##############################################  check currency  ########################################
		try:
			currency = page_soup.findAll("h2",{"class":"box-headline"})
			currency = currency[0].text.split('(in ')
			currency = currency[1]
			currency = currency.replace(')','')
			currency_index = 'D' + str(i)
			rawdata_worksheet[currency_index] = currency
		except Exception as e:
			print ('no currency')


		###################################### Boerse.de      ##############################################

		#########################################      grab price Feb 2009 ####################################

		try:
			browser.get('http://www.boerse.de/historische-kurse/wertpapier/'+ isin + '_jahr,2009#jahr')
			for x in range(1,13):
				price = browser.find_element_by_css_selector('table.table:nth-child(5) > tbody:nth-child(1) > tr:nth-child('+ str(x+1) + ') > td:nth-child(5)').text
				rawdata_worksheet.cell(row = i, column = x + 4, value = price)
		except Exception as e:
			print('no 2009 data')


		#########################################      grab price Feb 2010 ####################################

		try:
			browser.get('http://www.boerse.de/historische-kurse/wertpapier/'+ isin + '_jahr,2010#jahr')
			for x in range(1,13):
				price = browser.find_element_by_css_selector('table.table:nth-child(5) > tbody:nth-child(1) > tr:nth-child('+ str(x+1) +') > td:nth-child(5)').text
				rawdata_worksheet.cell(row = i, column = x + 16, value = price)
		except Exception as e:
			print('no 2010 data')


		#########################################      grab price Feb 2011 ####################################

		try:
			browser.get('http://www.boerse.de/historische-kurse/wertpapier/'+ isin + '_jahr,2011#jahr')
			for x in range(1,13):
				price = browser.find_element_by_css_selector('table.table:nth-child(5) > tbody:nth-child(1) > tr:nth-child(' + str(x+1) +') > td:nth-child(5)').text
				rawdata_worksheet.cell(row = i, column = x + 28, value = price)
		except Exception as e:
			print('no 2011 data')


		#########################################      grab price Feb 2012 ####################################

		try:
			browser.get('http://www.boerse.de/historische-kurse/wertpapier/'+ isin + '_jahr,2012#jahr')
			for x in range(1,13):
				price = browser.find_element_by_css_selector('table.table:nth-child(5) > tbody:nth-child(1) > tr:nth-child(' + str(x+1) +') > td:nth-child(5)').text
				rawdata_worksheet.cell(row = i, column = x + 40, value = price)
		except Exception as e:
			print('no 2012 data')


		#########################################      grab price Feb 2012 ####################################

		try:
			browser.get('http://www.boerse.de/historische-kurse/wertpapier/'+ isin + '_jahr,2013#jahr')
			for x in range(1,13):
				price = browser.find_element_by_css_selector('table.table:nth-child(5) > tbody:nth-child(1) > tr:nth-child(' + str(x+1) +') > td:nth-child(5)').text
				rawdata_worksheet.cell(row = i, column = x + 52, value = price)
		except Exception as e:
			print('no 2013 data')

		#########################################      grab price Feb 2012 ####################################

		try:
			browser.get('http://www.boerse.de/historische-kurse/wertpapier/'+ isin + '_jahr,2014#jahr')
			for x in range(1,13):
				price = browser.find_element_by_css_selector('table.table:nth-child(5) > tbody:nth-child(1) > tr:nth-child(' + str(x+1) +') > td:nth-child(5)').text
				rawdata_worksheet.cell(row = i, column = x + 64, value = price)
		except Exception as e:
			print('no 2014 data')


		#########################################      grab price Feb 2012 ####################################

		try:
			browser.get('http://www.boerse.de/historische-kurse/wertpapier/'+ isin + '_jahr,2015#jahr')
			for x in range(1,13):
				price = browser.find_element_by_css_selector('table.table:nth-child(5) > tbody:nth-child(1) > tr:nth-child(' + str(x+1) +') > td:nth-child(5)').text
				rawdata_worksheet.cell(row = i, column = x + 76, value = price)
		except Exception as e:
			print('no 2015 data')

				#########################################      grab price Feb 2012 ####################################

		try:
			browser.get('http://www.boerse.de/historische-kurse/wertpapier/'+ isin + '_jahr,2016#jahr')
			for x in range(1,13):
				price = browser.find_element_by_css_selector('table.table:nth-child(5) > tbody:nth-child(1) > tr:nth-child(' + str(x+1) + ') > td:nth-child(5)').text
				rawdata_worksheet.cell(row = i, column = x + 88, value = price)
		except Exception as e:
			print('no 2016 data')


		#########################################      grab price 2017 ######################################
		try:
			browser.get('http://www.boerse.de/historische-kurse/wertpapier/'+ isin + '_jahr,2017#jahr')
			for x in range(1,9):
				price = browser.find_element_by_css_selector('table.table:nth-child(5) > tbody:nth-child(1) > tr:nth-child(' + str(x+1) + ') > td:nth-child(5)').text
				rawdata_worksheet.cell(row = i, column = x + 96, value = price)
		except Exception as e:
			print('no 2017 data')

		rawdata_workbook.save('History.xlsx')

		browser.close()

	except Exception as e:
		print('error')


